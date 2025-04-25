"""
Excel Format Handler for legacy data conversion

This handler provides support for Excel files (XLS, XLSX), which are
commonly used for property assessment data sharing and reporting.
"""

import os
import logging
import numpy as np
import pandas as pd
from typing import Dict, List, Any, Tuple, Optional, Union, Generator
from legacy_converter import FormatHandler, ColumnMapping

# Configure logging
logger = logging.getLogger(__name__)

class ExcelHandler(FormatHandler):
    """Handler for Excel files (XLS, XLSX)"""
    
    def __init__(self):
        """Initialize Excel handler"""
        self.sheet_name = 0  # Default to first sheet
    
    def can_handle(self, file_path: str) -> bool:
        """
        Check if this handler can process the given file
        
        Args:
            file_path: Path to the file
            
        Returns:
            True if this handler can process the file, False otherwise
        """
        # Check extension
        _, ext = os.path.splitext(file_path.lower())
        
        if ext in [".xls", ".xlsx", ".xlsm"]:
            return True
        
        # Check file header for Excel signatures
        try:
            with open(file_path, "rb") as f:
                header = f.read(8)
                
                # Check for Excel file signatures
                # D0 CF 11 E0 A1 B1 1A E1 = OLE2 format (XLS)
                # 50 4B 03 04 = ZIP format (XLSX)
                if header.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1") or header.startswith(b"PK\x03\x04"):
                    return True
        except:
            pass
        
        return False
    
    def read_schema(self, file_path: str) -> Dict[str, Any]:
        """
        Read the schema of the Excel file
        
        Args:
            file_path: Path to the file
            
        Returns:
            Dict with schema information
        """
        try:
            # Get list of sheet names
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names
            
            # If multiple sheets, use the first non-empty one or the first one
            if len(sheet_names) > 1:
                for sheet in sheet_names:
                    # Read a sample to check if sheet is empty
                    sample = pd.read_excel(file_path, sheet_name=sheet, nrows=5)
                    if not sample.empty:
                        self.sheet_name = sheet
                        break
                else:
                    # If all sheets are empty, use the first one
                    self.sheet_name = sheet_names[0]
            else:
                self.sheet_name = sheet_names[0]
            
            # Read a sample of the data
            df = pd.read_excel(file_path, sheet_name=self.sheet_name, nrows=100)
            
            # Collect column information
            columns = {}
            for i, col_name in enumerate(df.columns):
                # Get dtype and convert to string representation
                dtype = df[col_name].dtype
                data_type = self._pandas_dtype_to_string(dtype)
                
                # Get sample values
                sample_values = df[col_name].head(5).tolist()
                
                # Convert numpy values to Python types for serialization
                sample_values = [self._numpy_to_python(val) for val in sample_values]
                
                columns[str(col_name)] = {
                    "index": i,
                    "name": str(col_name),
                    "type": data_type,
                    "nullable": df[col_name].isna().any(),
                    "description": "",
                    "sample_values": sample_values
                }
            
            # Estimate row count
            row_count_estimate = self._estimate_row_count(file_path, self.sheet_name)
            
            return {
                "columns": columns,
                "file_size": os.path.getsize(file_path),
                "sheet_name": self.sheet_name,
                "available_sheets": sheet_names,
                "row_count_estimate": row_count_estimate
            }
        except Exception as e:
            logger.error(f"Error reading Excel schema: {str(e)}")
            return {"error": str(e)}
    
    def read_data(self, file_path: str, batch_size: int = 1000) -> Generator[pd.DataFrame, None, None]:
        """
        Read data from the Excel file as a generator
        
        Args:
            file_path: Path to the file
            batch_size: Number of rows to read at a time
            
        Returns:
            Generator yielding DataFrames
        """
        try:
            # Get total row count estimate
            row_count = self._estimate_row_count(file_path, self.sheet_name)
            
            # Process in batches
            for start_row in range(0, row_count, batch_size):
                # Read batch from Excel file
                df = pd.read_excel(
                    file_path,
                    sheet_name=self.sheet_name,
                    skiprows=start_row if start_row == 0 else start_row + 1,  # Add 1 to skip header except for first batch
                    nrows=batch_size,
                    header=0 if start_row == 0 else None  # Only use header for first batch
                )
                
                # If not the first batch, set column names from first batch
                if start_row > 0:
                    first_batch = pd.read_excel(file_path, sheet_name=self.sheet_name, nrows=1)
                    df.columns = first_batch.columns
                
                # Stop if batch is empty
                if df.empty:
                    break
                
                # Clean column names
                df.columns = [str(col).strip() if isinstance(col, str) else str(col) for col in df.columns]
                
                yield df
        except Exception as e:
            logger.error(f"Error reading Excel data: {str(e)}")
            
            # Fallback to single batch read
            try:
                logger.info("Trying alternative Excel reading method (single batch)")
                df = pd.read_excel(file_path, sheet_name=self.sheet_name)
                
                # Clean column names
                df.columns = [str(col).strip() if isinstance(col, str) else str(col) for col in df.columns]
                
                yield df
            except Exception as e2:
                logger.error(f"Error with alternative Excel reading: {str(e2)}")
                yield pd.DataFrame()
    
    def validate_mapping(self, file_path: str, mapping: List[ColumnMapping]) -> List[Dict[str, Any]]:
        """
        Validate a column mapping against the file
        
        Args:
            file_path: Path to the file
            mapping: List of column mappings
            
        Returns:
            List of validation issues
        """
        issues = []
        
        try:
            # Read schema for column information
            schema = self.read_schema(file_path)
            if "error" in schema:
                issues.append({
                    "type": "schema_error",
                    "message": f"Error reading schema: {schema['error']}"
                })
                return issues
            
            columns = schema.get("columns", {})
            
            # Check each mapping
            for i, col_map in enumerate(mapping):
                # Verify source column exists
                if col_map.source_column not in columns:
                    issues.append({
                        "type": "missing_source_column",
                        "mapping_index": i,
                        "source_column": col_map.source_column,
                        "message": f"Source column '{col_map.source_column}' not found in file"
                    })
                    continue
                
                # Check data type compatibility
                source_type = columns[col_map.source_column]["type"]
                target_type = col_map.data_type
                
                if not self._are_types_compatible(source_type, target_type):
                    issues.append({
                        "type": "type_mismatch",
                        "mapping_index": i,
                        "source_column": col_map.source_column,
                        "target_column": col_map.target_column,
                        "source_type": source_type,
                        "target_type": target_type,
                        "message": f"Source type '{source_type}' may not be compatible with target type '{target_type}'"
                    })
            
            return issues
        except Exception as e:
            logger.error(f"Error validating Excel mapping: {str(e)}")
            issues.append({
                "type": "validation_error",
                "message": f"Error validating mapping: {str(e)}"
            })
            return issues
    
    def _estimate_row_count(self, file_path: str, sheet_name: str) -> int:
        """
        Estimate the number of rows in an Excel file sheet
        
        Args:
            file_path: Path to the file
            sheet_name: Name of the sheet
            
        Returns:
            Estimated row count
        """
        try:
            # For fast row count estimation
            from openpyxl import load_workbook
            
            # For XLSX files
            if file_path.lower().endswith(".xlsx") or file_path.lower().endswith(".xlsm"):
                # Load workbook with read_only=True for better performance with large files
                wb = load_workbook(filename=file_path, read_only=True)
                ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
                
                # Get dimensions
                return ws.max_row - 1  # Subtract 1 for header
            else:
                # For XLS files, read with pandas
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                return len(df)
        except Exception as e:
            logger.error(f"Error estimating Excel row count: {str(e)}")
            
            # Fallback method
            try:
                # Try reading the file and counting rows
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                return len(df)
            except:
                return 0
    
    def _pandas_dtype_to_string(self, dtype) -> str:
        """
        Convert pandas dtype to string representation
        
        Args:
            dtype: pandas dtype object
            
        Returns:
            String representation of the data type
        """
        dtype_str = str(dtype)
        
        if "int" in dtype_str:
            return "integer"
        elif "float" in dtype_str:
            return "float"
        elif "datetime" in dtype_str:
            return "date"
        elif "bool" in dtype_str:
            return "boolean"
        else:
            return "string"
    
    def _numpy_to_python(self, val):
        """
        Convert numpy value to Python native type for serialization
        
        Args:
            val: Numpy value
            
        Returns:
            Python native value
        """
        if pd.isna(val):
            return None
        elif isinstance(val, np.integer):
            return int(val)
        elif isinstance(val, np.floating):
            return float(val)
        elif isinstance(val, np.bool_):
            return bool(val)
        elif isinstance(val, pd.Timestamp):
            return val.isoformat()
        else:
            return val
    
    def _are_types_compatible(self, source_type: str, target_type: str) -> bool:
        """
        Check if source and target types are compatible
        
        Args:
            source_type: Source data type
            target_type: Target data type
            
        Returns:
            True if types are compatible, False otherwise
        """
        # Define compatibility rules
        compatibility = {
            "string": ["string", "text"],
            "integer": ["integer", "float", "string", "text"],
            "float": ["float", "string", "text"],
            "date": ["date", "datetime", "string", "text"],
            "boolean": ["boolean", "integer", "string", "text"],
            "text": ["string", "text"]
        }
        
        return target_type in compatibility.get(source_type, [])