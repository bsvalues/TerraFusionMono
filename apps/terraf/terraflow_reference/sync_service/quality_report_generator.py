"""
Quality Report Generator

This module provides functionality for generating data quality reports in various formats,
including PDF and Excel.
"""

import os
import io
import uuid
import logging
import datetime
from typing import Dict, List, Any, Optional, Tuple
from sqlalchemy import func, desc, and_
from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app import app, db
from sync_service.models.data_quality import (
    DataQualityReport, DataAnomaly, DataQualityIssue, DataQualityRule
)

# Configure logging
logger = logging.getLogger(__name__)

class QualityReportGenerator:
    """
    Generator for data quality reports in PDF and Excel formats.
    """
    
    def __init__(self):
        """Initialize the report generator."""
        self.version = "1.0"
    
    def generate_pdf_report(self, report_id: Optional[int] = None, 
                          start_date: Optional[datetime.datetime] = None,
                          end_date: Optional[datetime.datetime] = None,
                          save_to_db: bool = True,
                          options: Optional[Dict[str, Any]] = None) -> Tuple[bytes, str, Optional[int]]:
        """
        Generate a PDF report for the specified time period or report ID.
        
        Args:
            report_id: Optional ID of a specific report to use as the base
            start_date: Optional start date for the report period
            end_date: Optional end date for the report period
            save_to_db: Whether to save the report metadata to the database
            options: Optional dictionary of report generation options like include_anomalies,
                    include_issues, include_recommendations
            
        Returns:
            Tuple of (PDF bytes, filename, report_id)
        """
        # Set default options if none provided
        if options is None:
            options = {
                'include_anomalies': True,
                'include_issues': True,
                'include_recommendations': True
            }
            
        # Get report data
        summary, anomaly_summary, recent_anomalies, table_metrics, recommendations = self._get_report_data(
            report_id, start_date, end_date, options
        )
        
        # Generate report ID and date
        report_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        report_uuid = str(uuid.uuid4())[:8]
        
        # Render the HTML template
        rendered_html = self._render_html_template(
            report_date=report_date,
            report_id=report_uuid,
            version=self.version,
            summary=summary,
            anomaly_summary=anomaly_summary,
            recent_anomalies=recent_anomalies,
            table_metrics=table_metrics,
            recommendations=recommendations,
            options=options
        )
        
        # Convert HTML to PDF
        pdf_bytes = self._html_to_pdf(rendered_html)
        
        # Generate filename with date and ID
        filename = f"quality_report_{datetime.datetime.now().strftime('%Y%m%d')}_{report_uuid}.pdf"
        
        # Save report to filesystem
        reports_dir = os.path.join('uploads', 'reports')
        if not os.path.exists(reports_dir):
            os.makedirs(reports_dir, exist_ok=True)
            
        file_path = os.path.join(reports_dir, filename)
        with open(file_path, 'wb') as f:
            f.write(pdf_bytes)
            
        # Save report metadata to database
        db_report_id = None
        if save_to_db:
            try:
                # Create report record
                report = DataQualityReport(
                    report_name=f"Quality Report {datetime.datetime.now().strftime('%Y-%m-%d')}",
                    report_type='pdf',
                    tables_checked=summary.get('tables_checked', []),
                    overall_score=summary.get('overall_score', 0),
                    report_data={
                        'summary': summary,
                        'anomaly_summary': anomaly_summary,
                        'recent_anomalies': recent_anomalies,
                        'table_metrics': table_metrics,
                        'recommendations': recommendations
                    },
                    critical_issues=summary.get('critical_issues', 0),
                    high_issues=summary.get('high_issues', 0),
                    medium_issues=summary.get('medium_issues', 0),
                    low_issues=summary.get('low_issues', 0),
                    report_file_path=file_path,
                    report_format='pdf',
                    start_date=start_date,
                    end_date=end_date,
                    created_at=datetime.datetime.now()
                )
                
                db.session.add(report)
                db.session.commit()
                db_report_id = report.id
                
                logger.info(f"Report saved to database with ID: {db_report_id}")
            except Exception as e:
                logger.error(f"Error saving report to database: {str(e)}")
                db.session.rollback()
        
        return pdf_bytes, filename, db_report_id
    
    def _get_report_data(self, report_id: Optional[int] = None,
                        start_date: Optional[datetime.datetime] = None,
                        end_date: Optional[datetime.datetime] = None,
                        options: Optional[Dict[str, Any]] = None) -> Tuple[Dict, List, List, List, List]:
        """
        Get data for the report from the database.
        
        Args:
            report_id: Optional ID of a specific report to use as the base
            start_date: Optional start date for the report period
            end_date: Optional end date for the report period
            
        Returns:
            Tuple of (summary, anomaly_summary, recent_anomalies, table_metrics, recommendations)
        """
        if end_date is None:
            end_date = datetime.datetime.now()
            
        if start_date is None:
            # Default to one month before end date
            start_date = end_date - datetime.timedelta(days=30)
            
        # Get the latest report if no report_id is provided
        latest_report = None
        if report_id:
            latest_report = DataQualityReport.query.get(report_id)
        else:
            latest_report = DataQualityReport.query.order_by(DataQualityReport.created_at.desc()).first()
            
        # Get previous report for trend calculation
        prev_report = None
        if latest_report:
            prev_report = DataQualityReport.query.filter(
                DataQualityReport.created_at < latest_report.created_at
            ).order_by(DataQualityReport.created_at.desc()).first()
        
        # Create summary data
        summary = self._create_summary(latest_report, prev_report)
        
        # Set default options if not provided
        if options is None:
            options = {
                'include_anomalies': True,
                'include_issues': True,
                'include_recommendations': True
            }

        # Get anomaly summary data grouped by table (if requested)
        anomaly_summary = []
        if options.get('include_anomalies', True):
            anomaly_summary = self._get_anomaly_summary(start_date, end_date)
        
        # Get recent anomalies (if requested)
        recent_anomalies = []
        if options.get('include_anomalies', True):
            recent_anomalies = self._get_recent_anomalies(start_date, end_date)
        
        # Get table-specific metrics
        table_metrics = self._get_table_metrics(latest_report)
        
        # Generate recommendations based on the data (if requested)
        recommendations = []
        if options.get('include_recommendations', True):
            recommendations = self._generate_recommendations(
                summary, anomaly_summary, recent_anomalies, table_metrics
            )
        
        return summary, anomaly_summary, recent_anomalies, table_metrics, recommendations
    
    def _create_summary(self, report: Optional[DataQualityReport], prev_report: Optional[DataQualityReport]) -> Dict:
        """
        Create a summary dictionary from the report data with trend indicators.
        
        Args:
            report: The current report
            prev_report: The previous report for trend comparison
            
        Returns:
            Summary dictionary with metrics and trends
        """
        summary = {
            'overall_score': 0,
            'critical_issues': 0,
            'high_issues': 0,
            'medium_issues': 0,
            'low_issues': 0,
            'critical_trend': 0,
            'high_trend': 0,
            'medium_trend': 0,
            'low_trend': 0
        }
        
        if report:
            summary['overall_score'] = report.overall_score or 0
            summary['critical_issues'] = report.critical_issues or 0
            summary['high_issues'] = report.high_issues or 0
            summary['medium_issues'] = report.medium_issues or 0
            summary['low_issues'] = report.low_issues or 0
            
            # Calculate trends if previous report exists
            if prev_report:
                summary['critical_trend'] = summary['critical_issues'] - (prev_report.critical_issues or 0)
                summary['high_trend'] = summary['high_issues'] - (prev_report.high_issues or 0)
                summary['medium_trend'] = summary['medium_issues'] - (prev_report.medium_issues or 0)
                summary['low_trend'] = summary['low_issues'] - (prev_report.low_issues or 0)
        else:
            # If no report exists, calculate summary from current data
            anomalies = DataAnomaly.query.filter_by(status='open').all()
            issues = DataQualityIssue.query.filter_by(status='open').all()
            
            # Count by severity
            for anomaly in anomalies:
                if anomaly.severity == 'critical':
                    summary['critical_issues'] += 1
                elif anomaly.severity == 'high' or anomaly.severity == 'error':
                    summary['high_issues'] += 1
                elif anomaly.severity == 'medium' or anomaly.severity == 'warning':
                    summary['medium_issues'] += 1
                elif anomaly.severity == 'low' or anomaly.severity == 'info':
                    summary['low_issues'] += 1
                    
            for issue in issues:
                if issue.severity == 'critical':
                    summary['critical_issues'] += 1
                elif issue.severity == 'high' or issue.severity == 'error':
                    summary['high_issues'] += 1
                elif issue.severity == 'medium' or issue.severity == 'warning':
                    summary['medium_issues'] += 1
                elif issue.severity == 'low' or issue.severity == 'info':
                    summary['low_issues'] += 1
            
            # Calculate simple quality score based on issues
            total_issues = summary['critical_issues'] * 10 + summary['high_issues'] * 5 + \
                        summary['medium_issues'] * 2 + summary['low_issues']
            
            # Basic calculation - can be refined
            summary['overall_score'] = max(0, 100 - min(100, total_issues))
            
        return summary
    
    def _get_anomaly_summary(self, start_date: datetime.datetime, end_date: datetime.datetime) -> List[Dict]:
        """
        Get anomaly summary grouped by table.
        
        Args:
            start_date: Start date for the reporting period
            end_date: End date for the reporting period
            
        Returns:
            List of dictionaries with anomaly summaries by table
        """
        # Query database for anomalies grouped by table
        tables_with_anomalies = db.session.query(
            DataAnomaly.table_name,
            func.count(DataAnomaly.id).label('count')
        ).filter(
            DataAnomaly.detected_at.between(start_date, end_date)
        ).group_by(
            DataAnomaly.table_name
        ).all()
        
        results = []
        for table_name, count in tables_with_anomalies:
            # For each table, find the most common anomaly type
            most_common = db.session.query(
                DataAnomaly.anomaly_type,
                func.count(DataAnomaly.id).label('type_count')
            ).filter(
                DataAnomaly.table_name == table_name,
                DataAnomaly.detected_at.between(start_date, end_date)
            ).group_by(
                DataAnomaly.anomaly_type
            ).order_by(
                func.count(DataAnomaly.id).desc()
            ).first()
            
            # Count open vs. resolved
            open_count = db.session.query(func.count(DataAnomaly.id)).filter(
                DataAnomaly.table_name == table_name,
                DataAnomaly.status == 'open',
                DataAnomaly.detected_at.between(start_date, end_date)
            ).scalar()
            
            resolved_count = db.session.query(func.count(DataAnomaly.id)).filter(
                DataAnomaly.table_name == table_name,
                DataAnomaly.status == 'resolved',
                DataAnomaly.detected_at.between(start_date, end_date)
            ).scalar()
            
            most_common_type = most_common[0] if most_common else 'unknown'
            status = f"{open_count} open, {resolved_count} resolved"
            
            results.append({
                'table_name': table_name,
                'count': count,
                'most_common_type': most_common_type,
                'status': status
            })
            
        return results
    
    def _get_recent_anomalies(self, start_date: datetime.datetime, end_date: datetime.datetime, limit: int = 10) -> List[Dict]:
        """
        Get the most recent anomalies.
        
        Args:
            start_date: Start date for the reporting period
            end_date: End date for the reporting period
            limit: Maximum number of anomalies to return
            
        Returns:
            List of recent anomalies
        """
        recent = DataAnomaly.query.filter(
            DataAnomaly.detected_at.between(start_date, end_date)
        ).order_by(
            DataAnomaly.detected_at.desc()
        ).limit(limit).all()
        
        result = []
        for anomaly in recent:
            result.append({
                'id': anomaly.id,
                'table_name': anomaly.table_name,
                'field_name': anomaly.field_name or 'N/A',
                'anomaly_type': anomaly.anomaly_type,
                'severity': anomaly.severity,
                'detected_at': anomaly.detected_at.strftime('%Y-%m-%d %H:%M')
            })
            
        return result
    
    def _get_table_metrics(self, report: Optional[DataQualityReport]) -> List[Dict]:
        """
        Get metrics by table from the report data.
        
        Args:
            report: The quality report
            
        Returns:
            List of table metrics
        """
        table_metrics = []
        
        # If we have a report with this data, use it
        if report and hasattr(report, 'report_data') and isinstance(report.report_data, dict):
            table_data = report.report_data.get('tables', {})
            for table_name, metrics in table_data.items():
                if isinstance(metrics, dict):
                    table_metrics.append({
                        'name': table_name,
                        'completeness': metrics.get('completeness', 0) * 100,
                        'accuracy': metrics.get('accuracy', 0) * 100,
                        'consistency': metrics.get('consistency', 0) * 100,
                        'overall': metrics.get('overall', 0) * 100
                    })
        else:
            # Otherwise, calculate metrics from current data
            # Get list of tables with rules or anomalies
            tables_with_rules = db.session.query(DataQualityRule.table_name).distinct().all()
            tables_with_anomalies = db.session.query(DataAnomaly.table_name).distinct().all()
            
            # Combine tables from both sources
            table_names = set([t[0] for t in tables_with_rules] + [t[0] for t in tables_with_anomalies])
            
            for table_name in table_names:
                # Count rules for this table
                rule_count = db.session.query(func.count(DataQualityRule.id)).filter(
                    DataQualityRule.table_name == table_name
                ).scalar()
                
                # Count open anomalies for this table
                anomaly_count = db.session.query(func.count(DataAnomaly.id)).filter(
                    DataAnomaly.table_name == table_name,
                    DataAnomaly.status == 'open'
                ).scalar()
                
                # Count open issues for this table
                issue_count = db.session.query(func.count(DataQualityIssue.id)).filter(
                    DataQualityIssue.table_name == table_name,
                    DataQualityIssue.status == 'open'
                ).scalar()
                
                # Simple metric calculation
                if rule_count > 0:
                    # Completeness - assume 100% minus a percentage based on issues
                    completeness = 100.0 - min(100, issue_count * 5)
                    
                    # Accuracy - assume impacted by anomalies
                    accuracy = 100.0 - min(100, anomaly_count * 10)
                    
                    # Consistency - combination of the two
                    consistency = (completeness + accuracy) / 2
                    
                    # Overall
                    overall = (completeness + accuracy + consistency) / 3
                    
                    table_metrics.append({
                        'name': table_name,
                        'completeness': completeness,
                        'accuracy': accuracy,
                        'consistency': consistency, 
                        'overall': overall
                    })
        
        # Sort by overall score ascending (worst first)
        table_metrics.sort(key=lambda x: x['overall'])
        
        return table_metrics
    
    def _generate_recommendations(self, summary: Dict, anomaly_summary: List, 
                               recent_anomalies: List, table_metrics: List) -> List[str]:
        """
        Generate actionable recommendations based on the report data.
        
        Args:
            summary: Report summary data
            anomaly_summary: Anomaly summary by table
            recent_anomalies: List of recent anomalies
            table_metrics: Table-specific metrics
            
        Returns:
            List of recommendation strings
        """
        recommendations = []
        
        # Add recommendations based on the data
        
        # Critical issues always get a recommendation
        if summary['critical_issues'] > 0:
            recommendations.append(
                f"Address {summary['critical_issues']} critical issue(s) immediately to prevent data corruption or system failures."
            )
        
        # For tables with poor metrics
        for table in table_metrics[:3]:  # Top 3 worst tables
            if table['overall'] < 70:
                recommendations.append(
                    f"Improve data quality in the '{table['name']}' table (current score: {table['overall']:.1f}%) by addressing consistency and accuracy issues."
                )
                
        # For tables with many anomalies
        tables_with_many_anomalies = [t for t in anomaly_summary if t['count'] > 5]
        for table in tables_with_many_anomalies[:2]:  # Top 2 tables with most anomalies
            recommendations.append(
                f"Investigate unusual patterns in the '{table['table_name']}' table, which has {table['count']} anomalies (mostly {table['most_common_type']})."
            )
            
        # Add a general recommendation if the overall score is low
        if summary['overall_score'] < 80:
            recommendations.append(
                "Implement additional data validation rules to improve the overall data quality score."
            )
            
        # Make sure we have at least one recommendation
        if not recommendations:
            recommendations.append(
                "Continue monitoring data quality and maintain current validation practices."
            )
            
        return recommendations
    
    def _render_html_template(self, **kwargs) -> str:
        """
        Render the HTML template with the provided data.
        
        Args:
            **kwargs: Template variables
            
        Returns:
            Rendered HTML string
        """
        with app.app_context():
            rendered = app.jinja_env.get_template('reports/quality_report.html').render(**kwargs)
            return rendered
    
    def _html_to_pdf(self, html_content: str) -> bytes:
        """
        Convert HTML content to PDF using WeasyPrint.
        
        Args:
            html_content: HTML content to convert
            
        Returns:
            PDF as bytes
        """
        pdf = HTML(string=html_content).write_pdf()
        return pdf
        
    def generate_excel_report(self, report_id: Optional[int] = None, 
                             start_date: Optional[datetime.datetime] = None,
                             end_date: Optional[datetime.datetime] = None,
                             save_to_db: bool = True,
                             options: Optional[Dict[str, Any]] = None) -> Tuple[bytes, str, Optional[int]]:
        """
        Generate an Excel report for the specified time period or report ID.
        
        Args:
            report_id: Optional ID of a specific report to use as the base
            start_date: Optional start date for the report period
            end_date: Optional end date for the report period
            save_to_db: Whether to save the report metadata to the database
            options: Optional dictionary of report generation options like include_anomalies,
                    include_issues, include_recommendations
            
        Returns:
            Tuple of (Excel bytes, filename, report_id)
        """
        # Set default options if none provided
        if options is None:
            options = {
                'include_anomalies': True,
                'include_issues': True,
                'include_recommendations': True
            }
            
        # Get report data
        summary, anomaly_summary, recent_anomalies, table_metrics, recommendations = self._get_report_data(
            report_id, start_date, end_date, options
        )
        
        # Generate report ID and date
        report_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        report_uuid = str(uuid.uuid4())[:8]
        
        # Create Excel workbook
        excel_bytes = self._generate_excel_workbook(
            report_date=report_date,
            report_id=report_uuid,
            version=self.version,
            summary=summary,
            anomaly_summary=anomaly_summary,
            recent_anomalies=recent_anomalies,
            table_metrics=table_metrics,
            recommendations=recommendations,
            options=options
        )
        
        # Generate filename with date and ID
        filename = f"quality_report_{datetime.datetime.now().strftime('%Y%m%d')}_{report_uuid}.xlsx"
        
        # Save report to filesystem
        reports_dir = os.path.join('uploads', 'reports')
        if not os.path.exists(reports_dir):
            os.makedirs(reports_dir, exist_ok=True)
            
        file_path = os.path.join(reports_dir, filename)
        with open(file_path, 'wb') as f:
            f.write(excel_bytes)
            
        # Save report metadata to database
        db_report_id = None
        if save_to_db:
            try:
                # Create report record
                report = DataQualityReport(
                    report_name=f"Quality Report {datetime.datetime.now().strftime('%Y-%m-%d')}",
                    report_type='excel',
                    tables_checked=summary.get('tables_checked', []),
                    overall_score=summary.get('overall_score', 0),
                    report_data={
                        'summary': summary,
                        'anomaly_summary': anomaly_summary,
                        'recent_anomalies': recent_anomalies,
                        'table_metrics': table_metrics,
                        'recommendations': recommendations
                    },
                    critical_issues=summary.get('critical_issues', 0),
                    high_issues=summary.get('high_issues', 0),
                    medium_issues=summary.get('medium_issues', 0),
                    low_issues=summary.get('low_issues', 0),
                    report_file_path=file_path,
                    report_format='excel',
                    start_date=start_date,
                    end_date=end_date,
                    created_at=datetime.datetime.now()
                )
                
                db.session.add(report)
                db.session.commit()
                db_report_id = report.id
                
                logger.info(f"Excel report saved to database with ID: {db_report_id}")
            except Exception as e:
                logger.error(f"Error saving Excel report to database: {str(e)}")
                db.session.rollback()
        
        return excel_bytes, filename, db_report_id
    
    def _generate_excel_workbook(self, **kwargs) -> bytes:
        """
        Generate an Excel workbook with the provided data.
        
        Args:
            **kwargs: Template variables including report data
            
        Returns:
            Excel file as bytes
        """
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Remove the default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Add Summary sheet
        self._add_summary_sheet(wb, **kwargs)
        
        # Add Anomalies sheet if included
        if kwargs.get('options', {}).get('include_anomalies', True) and kwargs.get('recent_anomalies'):
            self._add_anomalies_sheet(wb, **kwargs)
            
            # Add Anomaly Summary sheet if there's data
            if kwargs.get('anomaly_summary'):
                self._add_anomaly_summary_sheet(wb, **kwargs)
        
        # Add Table Metrics sheet if there's data
        if kwargs.get('table_metrics'):
            self._add_table_metrics_sheet(wb, **kwargs)
        
        # Add Recommendations sheet if included and there's data
        if kwargs.get('options', {}).get('include_recommendations', True) and kwargs.get('recommendations'):
            self._add_recommendations_sheet(wb, **kwargs)
        
        # Save the workbook to a BytesIO object
        excel_file = io.BytesIO()
        wb.save(excel_file)
        
        # Get Excel bytes
        excel_file.seek(0)
        excel_bytes = excel_file.getvalue()
        excel_file.close()
        
        return excel_bytes
    
    def _add_summary_sheet(self, wb, **kwargs):
        """Add a summary sheet to the workbook."""
        summary = kwargs.get('summary', {})
        report_date = kwargs.get('report_date', datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        report_id = kwargs.get('report_id', '')
        version = kwargs.get('version', '1.0')
        
        # Create sheet
        ws = wb.create_sheet("Summary")
        
        # Define styles
        title_font = Font(name='Arial', size=14, bold=True)
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=11)
        
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Add report title and metadata
        ws['A1'] = "Data Quality Report"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Report Date: {report_date}"
        ws['A3'] = f"Report ID: {report_id}"
        ws['A4'] = f"Version: {version}"
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Add overall quality score
        ws['A6'] = "Overall Quality Score"
        ws['A6'].font = header_font
        ws.merge_cells('A6:E6')
        
        ws['A7'] = "Score:"
        ws['B7'] = summary.get('overall_score', 0)
        
        # Color score based on value
        score_cell = ws['B7']
        if summary.get('overall_score', 0) >= 90:
            score_cell.fill = good_fill
        elif summary.get('overall_score', 0) >= 70:
            score_cell.fill = warning_fill
        else:
            score_cell.fill = error_fill
            
        # Add issue summary
        ws['A9'] = "Issue Summary"
        ws['A9'].font = header_font
        ws.merge_cells('A9:E9')
        
        # Add headers
        headers = ["Severity", "Count", "Trend"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=10, column=i+1)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        # Add issue data
        issue_rows = [
            ("Critical", summary.get('critical_issues', 0), summary.get('critical_trend', 0)),
            ("High", summary.get('high_issues', 0), summary.get('high_trend', 0)),
            ("Medium", summary.get('medium_issues', 0), summary.get('medium_trend', 0)),
            ("Low", summary.get('low_issues', 0), summary.get('low_trend', 0))
        ]
        
        for i, (severity, count, trend) in enumerate(issue_rows):
            row = 11 + i
            
            # Severity
            ws.cell(row=row, column=1).value = severity
            
            # Count
            count_cell = ws.cell(row=row, column=2)
            count_cell.value = count
            
            # Apply fill color based on severity
            if severity == "Critical" and count > 0:
                count_cell.fill = error_fill
            elif severity == "High" and count > 0:
                count_cell.fill = warning_fill
            
            # Trend
            trend_cell = ws.cell(row=row, column=3)
            if trend != 0:
                trend_prefix = "+" if trend > 0 else ""
                trend_cell.value = f"{trend_prefix}{trend}"
                
                # Color code trends
                if trend > 0:
                    trend_cell.fill = error_fill
                else:
                    trend_cell.fill = good_fill
    
    def _add_anomalies_sheet(self, wb, **kwargs):
        """Add the recent anomalies sheet to the workbook."""
        recent_anomalies = kwargs.get('recent_anomalies', [])
        
        # Create sheet
        ws = wb.create_sheet("Recent Anomalies")
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        
        # Add headers
        headers = ["ID", "Table", "Field", "Anomaly Type", "Severity", "Detected At"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=1, column=i+1)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        # Add anomaly data
        for i, anomaly in enumerate(recent_anomalies):
            row = i + 2
            
            ws.cell(row=row, column=1).value = anomaly.get('id', '')
            ws.cell(row=row, column=2).value = anomaly.get('table_name', '')
            ws.cell(row=row, column=3).value = anomaly.get('field_name', '')
            ws.cell(row=row, column=4).value = anomaly.get('anomaly_type', '')
            ws.cell(row=row, column=5).value = anomaly.get('severity', '')
            ws.cell(row=row, column=6).value = anomaly.get('detected_at', '')
            
            # Apply fill color based on severity
            severity_cell = ws.cell(row=row, column=5)
            if anomaly.get('severity') == 'critical':
                severity_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif anomaly.get('severity') in ['high', 'error']:
                severity_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    def _add_anomaly_summary_sheet(self, wb, **kwargs):
        """Add the anomaly summary sheet to the workbook."""
        anomaly_summary = kwargs.get('anomaly_summary', [])
        
        # Create sheet
        ws = wb.create_sheet("Anomaly Summary")
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        
        # Add headers
        headers = ["Table", "Anomaly Count", "Most Common Type", "Status"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=1, column=i+1)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        # Add summary data
        for i, summary_item in enumerate(anomaly_summary):
            row = i + 2
            
            ws.cell(row=row, column=1).value = summary_item.get('table_name', '')
            ws.cell(row=row, column=2).value = summary_item.get('count', 0)
            ws.cell(row=row, column=3).value = summary_item.get('most_common_type', '')
            ws.cell(row=row, column=4).value = summary_item.get('status', '')
    
    def _add_table_metrics_sheet(self, wb, **kwargs):
        """Add the table metrics sheet to the workbook."""
        table_metrics = kwargs.get('table_metrics', [])
        
        # Create sheet
        ws = wb.create_sheet("Table Metrics")
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Add headers
        headers = ["Table", "Completeness %", "Accuracy %", "Consistency %", "Overall %"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=1, column=i+1)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        # Add metric data
        for i, metric in enumerate(table_metrics):
            row = i + 2
            
            ws.cell(row=row, column=1).value = metric.get('name', '')
            ws.cell(row=row, column=2).value = metric.get('completeness', 0)
            ws.cell(row=row, column=3).value = metric.get('accuracy', 0)
            ws.cell(row=row, column=4).value = metric.get('consistency', 0)
            ws.cell(row=row, column=5).value = metric.get('overall', 0)
            
            # Apply conditional formatting to metric values
            for col in range(2, 6):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                
                if value >= 90:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value >= 70:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def _add_recommendations_sheet(self, wb, **kwargs):
        """Add the recommendations sheet to the workbook."""
        recommendations = kwargs.get('recommendations', [])
        
        # Create sheet
        ws = wb.create_sheet("Recommendations")
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 65
        
        # Add title
        ws['A1'] = "Data Quality Recommendations"
        ws['A1'].font = header_font
        ws.merge_cells('A1:B1')
        
        # Add recommendations
        for i, recommendation in enumerate(recommendations):
            row = i + 3
            
            ws.cell(row=row, column=1).value = f"Recommendation {i+1}:"
            ws.cell(row=row, column=2).value = recommendation
            
            # Apply wrap text to long recommendations
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)


# Initialize the report generator
report_generator = QualityReportGenerator()