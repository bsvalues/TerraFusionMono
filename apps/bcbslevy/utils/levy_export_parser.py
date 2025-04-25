"""
Levy Export Parser for the SaaS Levy Calculation System.

This module provides classes for parsing levy export files in various formats,
including text (.txt), Excel (.xls, .xlsx), and XML formats.
"""

import os
import re
import csv
import json
import logging
import tempfile
import io
from enum import Enum
from typing import List, Dict, Any, Union, Optional, Tuple
from datetime import datetime
from pathlib import Path

# Third-party libraries
import openpyxl
import pandas as pd
import xlrd

# Setup logging
logger = logging.getLogger(__name__)


class LevyExportFormat(Enum):
    """Enumeration of supported levy export file formats."""
    TXT = 'txt'
    XLS = 'xls'
    XLSX = 'xlsx'
    XML = 'xml'
    CSV = 'csv'
    JSON = 'json'
    UNKNOWN = 'unknown'


class LevyRecord:
    """A single levy record extracted from a levy export file."""
    
    def __init__(self, data: Dict[str, Any]):
        """
        Initialize a levy record with the provided data.
        
        Args:
            data: Dictionary of levy record data
        """
        self.data = data
        
    def __getitem__(self, key: str) -> Any:
        """
        Get a value from the record data.
        
        Args:
            key: The key to look up
            
        Returns:
            The value for the key or None if not found
        """
        return self.data.get(key)
    
    def get(self, key: str, default: Any = None) -> Any:
        """
        Get a value from the record data with a default.
        
        Args:
            key: The key to look up
            default: Default value if key not found
            
        Returns:
            The value for the key or the default if not found
        """
        return self.data.get(key, default)


class LevyExportData:
    """Container for levy export data extracted from a file."""
    
    def __init__(self, records: List[Dict[str, Any]], metadata: Optional[Dict[str, Any]] = None):
        """
        Initialize with records and optional metadata.
        
        Args:
            records: List of levy record dictionaries
            metadata: Optional metadata about the export
        """
        self.records = [LevyRecord(record) for record in records]
        self.metadata = metadata or {}
        
    def __len__(self) -> int:
        """Get the number of records."""
        return len(self.records)
    
    def get_years(self) -> List[int]:
        """
        Get a list of all years in the data.
        
        Returns:
            List of years found in the records
        """
        years = set()
        for record in self.records:
            year = record['year']
            if year:
                try:
                    years.add(int(year))
                except (ValueError, TypeError):
                    pass
        return sorted(list(years))
    
    def get_tax_districts(self) -> List[str]:
        """
        Get a list of all tax districts in the data.
        
        Returns:
            List of unique tax district IDs
        """
        districts = set()
        for record in self.records:
            district = record['tax_district_id']
            if district:
                districts.add(str(district))
        return sorted(list(districts))
    
    def get_levy_codes(self) -> List[str]:
        """
        Get a list of all levy codes in the data.
        
        Returns:
            List of unique levy codes
        """
        codes = set()
        for record in self.records:
            code = record['levy_cd']
            if code:
                codes.add(str(code))
        return sorted(list(codes))


class LevyExportParser:
    """Parser for levy export files in various formats."""
    
    @classmethod
    def detect_format(cls, file_path: Union[str, Path]) -> LevyExportFormat:
        """
        Detect the format of a levy export file.
        
        Args:
            file_path: Path to the file
            
        Returns:
            The detected file format as a LevyExportFormat enum
        """
        file_path = Path(file_path)
        extension = file_path.suffix.lower().lstrip('.')
        
        if extension == 'txt':
            # Check if it's actually a CSV file
            try:
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    first_line = f.readline().strip()
                    if ',' in first_line and len(first_line.split(',')) >= 3:
                        # If it has commas and at least 3 columns, treat as CSV
                        logger.info(f"File has .txt extension but detected as CSV format: {file_path}")
                        return LevyExportFormat.CSV
            except Exception as e:
                logger.warning(f"Error checking if TXT file is CSV: {str(e)}")
            
            return LevyExportFormat.TXT
        elif extension == 'xls':
            return LevyExportFormat.XLS
        elif extension == 'xlsx':
            return LevyExportFormat.XLSX
        elif extension == 'xml':
            return LevyExportFormat.XML
        elif extension == 'csv':
            return LevyExportFormat.CSV
        elif extension == 'json':
            return LevyExportFormat.JSON
        else:
            # Try to detect by content
            try:
                with open(file_path, 'rb') as f:
                    header = f.read(4)
                    if header == b'PK\x03\x04':  # XLSX files start with this signature
                        return LevyExportFormat.XLSX
                    if header[:2] == b'\xd0\xcf':  # XLS files start with this signature
                        return LevyExportFormat.XLS
            except Exception as e:
                logger.warning(f"Error detecting file format by content: {str(e)}")
            
            return LevyExportFormat.UNKNOWN
    
    @classmethod
    def parse_file(cls, file_path: Union[str, Path]) -> LevyExportData:
        """
        Parse a levy export file.
        
        Args:
            file_path: Path to the file
            
        Returns:
            LevyExportData object containing the parsed data
            
        Raises:
            ValueError: If the file format is not supported
        """
        file_path = Path(file_path)
        format = cls.detect_format(file_path)
        
        if format == LevyExportFormat.TXT:
            return cls._parse_txt(file_path)
        elif format == LevyExportFormat.XLS:
            return cls._parse_xls(file_path)
        elif format == LevyExportFormat.XLSX:
            return cls._parse_xlsx(file_path)
        elif format == LevyExportFormat.XML:
            return cls._parse_xml(file_path)
        elif format == LevyExportFormat.CSV:
            return cls._parse_csv(file_path)
        elif format == LevyExportFormat.JSON:
            return cls._parse_json(file_path)
        else:
            raise ValueError(f"Unsupported file format: {format}")
    
    @classmethod
    def _parse_txt(cls, file_path: Path) -> LevyExportData:
        """
        Parse a text format levy export file.
        
        Args:
            file_path: Path to the TXT file
            
        Returns:
            LevyExportData object containing the parsed data
        """
        logger.info(f"Parsing TXT levy export file: {file_path}")
        
        records = []
        year = datetime.now().year
        header_pattern = re.compile(r'^\s*LEVY CODE\s+RATE\s+LEVY\s+VALUE\s*$', re.IGNORECASE)
        
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                lines = f.readlines()
                
                # Scan for headers and extract year if available
                for i, line in enumerate(lines[:20]):  # Check first 20 lines for headers
                    if re.search(r'(LEVY|TAX)\s+YEAR\s*[:=\s]\s*(\d{4})', line, re.IGNORECASE):
                        year_match = re.search(r'(\d{4})', line)
                        if year_match:
                            year = int(year_match.group(1))
                            logger.info(f"Found year in header: {year}")
                    
                    if header_pattern.search(line):
                        # Found the column headers, start parsing from next line
                        data_lines = lines[i+1:]
                        break
                else:
                    # If no header found, assume all lines are data
                    data_lines = lines
                
                # Parse data lines
                for line in data_lines:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    
                    # Try to parse fixed-width format
                    match = re.match(r'^\s*(\S+)\s+(\d*\.?\d*)\s+(\d*\.?\d*)\s+(\d*\.?\d*)', line)
                    if match:
                        levy_cd, rate, levy, value = match.groups()
                        
                        # Clean up and convert values
                        try:
                            rate = float(rate) if rate else None
                            levy = float(levy) if levy else None
                            value = float(value) if value else None
                        except ValueError:
                            logger.warning(f"Failed to parse values from line: {line}")
                            continue
                        
                        # Check for linked levy code
                        levy_cd_linked = None
                        if '/' in levy_cd:
                            levy_cd, levy_cd_linked = levy_cd.split('/', 1)
                        
                        record = {
                            'tax_district_id': levy_cd,
                            'levy_cd': levy_cd,
                            'levy_cd_linked': levy_cd_linked,
                            'levy_rate': rate,
                            'levy_amount': levy,
                            'assessed_value': value,
                            'year': year,
                            'source': 'txt'
                        }
                        records.append(record)
        
        except Exception as e:
            logger.error(f"Error parsing TXT file {file_path}: {str(e)}")
            raise
        
        logger.info(f"Parsed {len(records)} records from TXT file")
        return LevyExportData(records, {'format': 'txt', 'year': year})
    
    @classmethod
    def _parse_xls(cls, file_path: Path) -> LevyExportData:
        """
        Parse an Excel .xls format levy export file.
        
        Args:
            file_path: Path to the XLS file
            
        Returns:
            LevyExportData object containing the parsed data
        """
        logger.info(f"Parsing XLS levy export file: {file_path}")
        
        try:
            # Open the workbook and select the first sheet
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
            
            # Attempt to find header row
            header_row = None
            year = datetime.now().year
            
            for row_idx in range(min(20, sheet.nrows)):  # Check first 20 rows
                row_values = [str(cell).strip().upper() for cell in sheet.row_values(row_idx)]
                row_text = ' '.join(row_values)
                
                # Look for year in header rows
                if re.search(r'(LEVY|TAX)\s+YEAR\s*[:=\s]\s*(\d{4})', row_text, re.IGNORECASE):
                    year_match = re.search(r'(\d{4})', row_text)
                    if year_match:
                        year = int(year_match.group(1))
                        logger.info(f"Found year in header: {year}")
                
                # Look for column headers
                if 'LEVY CODE' in row_text and 'RATE' in row_text:
                    header_row = row_idx
                    logger.info(f"Found header row at index {header_row}")
                    break
            
            # If header not found, try to infer it
            if header_row is None:
                logger.warning("No header row found, attempting to infer column structure")
                header_row = 0
            
            # Get column indices
            header_values = [str(cell).strip().upper() for cell in sheet.row_values(header_row)]
            
            col_indices = {
                'levy_cd': next((i for i, h in enumerate(header_values) if 'LEVY CODE' in h or 'CODE' in h), 0),
                'rate': next((i for i, h in enumerate(header_values) if 'RATE' in h), 1),
                'levy': next((i for i, h in enumerate(header_values) if 'LEVY' in h and 'CODE' not in h), 2),
                'value': next((i for i, h in enumerate(header_values) if 'VALUE' in h or 'ASSESSED' in h), 3)
            }
            
            # Parse data rows
            records = []
            for row_idx in range(header_row + 1, sheet.nrows):
                row_values = sheet.row_values(row_idx)
                
                if not row_values[col_indices['levy_cd']]:  # Skip empty rows
                    continue
                
                # Get values from appropriate columns
                levy_cd = str(row_values[col_indices['levy_cd']]).strip()
                
                # Handle potential float formatting issues
                try:
                    rate = float(row_values[col_indices['rate']]) if row_values[col_indices['rate']] else None
                except (ValueError, TypeError):
                    rate = None
                
                try:
                    levy = float(row_values[col_indices['levy']]) if row_values[col_indices['levy']] else None
                except (ValueError, TypeError):
                    levy = None
                
                try:
                    value = float(row_values[col_indices['value']]) if row_values[col_indices['value']] else None
                except (ValueError, TypeError):
                    value = None
                
                # Check for linked levy code
                levy_cd_linked = None
                if '/' in levy_cd:
                    levy_cd, levy_cd_linked = levy_cd.split('/', 1)
                
                record = {
                    'tax_district_id': levy_cd,
                    'levy_cd': levy_cd,
                    'levy_cd_linked': levy_cd_linked,
                    'levy_rate': rate,
                    'levy_amount': levy,
                    'assessed_value': value,
                    'year': year,
                    'source': 'xls'
                }
                records.append(record)
        
        except Exception as e:
            logger.error(f"Error parsing XLS file {file_path}: {str(e)}")
            raise
        
        logger.info(f"Parsed {len(records)} records from XLS file")
        return LevyExportData(records, {'format': 'xls', 'year': year})
    
    @classmethod
    def _parse_xlsx(cls, file_path: Path) -> LevyExportData:
        """
        Parse an Excel .xlsx format levy export file.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            LevyExportData object containing the parsed data
        """
        logger.info(f"Parsing XLSX levy export file: {file_path}")
        
        try:
            # Open the workbook and select the first worksheet
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Attempt to find header row
            header_row = None
            year = datetime.now().year
            
            for row_idx in range(1, min(21, ws.max_row + 1)):  # Check first 20 rows (1-indexed)
                row_values = [str(cell.value or '').strip().upper() for cell in ws[row_idx]]
                row_text = ' '.join(row_values)
                
                # Look for year in header rows
                if re.search(r'(LEVY|TAX)\s+YEAR\s*[:=\s]\s*(\d{4})', row_text, re.IGNORECASE):
                    year_match = re.search(r'(\d{4})', row_text)
                    if year_match:
                        year = int(year_match.group(1))
                        logger.info(f"Found year in header: {year}")
                
                # Look for column headers
                if 'LEVY CODE' in row_text and 'RATE' in row_text:
                    header_row = row_idx
                    logger.info(f"Found header row at index {header_row}")
                    break
            
            # If header not found, try to infer it
            if header_row is None:
                logger.warning("No header row found, attempting to infer column structure")
                header_row = 1
            
            # Get column indices
            header_values = [str(cell.value or '').strip().upper() for cell in ws[header_row]]
            
            col_indices = {
                'levy_cd': next((i for i, h in enumerate(header_values) if 'LEVY CODE' in h or 'CODE' in h), 0),
                'rate': next((i for i, h in enumerate(header_values) if 'RATE' in h), 1),
                'levy': next((i for i, h in enumerate(header_values) if 'LEVY' in h and 'CODE' not in h), 2),
                'value': next((i for i, h in enumerate(header_values) if 'VALUE' in h or 'ASSESSED' in h), 3)
            }
            
            # Parse data rows
            records = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                row_values = [cell.value for cell in ws[row_idx]]
                
                if not row_values or not row_values[col_indices['levy_cd']]:  # Skip empty rows
                    continue
                
                # Get values from appropriate columns
                levy_cd = str(row_values[col_indices['levy_cd']]).strip()
                
                # Handle potential formatting issues
                try:
                    rate = float(row_values[col_indices['rate']]) if row_values[col_indices['rate']] is not None else None
                except (ValueError, TypeError):
                    rate = None
                
                try:
                    levy = float(row_values[col_indices['levy']]) if row_values[col_indices['levy']] is not None else None
                except (ValueError, TypeError):
                    levy = None
                
                try:
                    value = float(row_values[col_indices['value']]) if row_values[col_indices['value']] is not None else None
                except (ValueError, TypeError):
                    value = None
                
                # Check for linked levy code
                levy_cd_linked = None
                if '/' in levy_cd:
                    levy_cd, levy_cd_linked = levy_cd.split('/', 1)
                
                record = {
                    'tax_district_id': levy_cd,
                    'levy_cd': levy_cd,
                    'levy_cd_linked': levy_cd_linked,
                    'levy_rate': rate,
                    'levy_amount': levy,
                    'assessed_value': value,
                    'year': year,
                    'source': 'xlsx'
                }
                records.append(record)
        
        except Exception as e:
            logger.error(f"Error parsing XLSX file {file_path}: {str(e)}")
            raise
        
        logger.info(f"Parsed {len(records)} records from XLSX file")
        return LevyExportData(records, {'format': 'xlsx', 'year': year})
    
    @classmethod
    def _parse_csv(cls, file_path: Path) -> LevyExportData:
        """
        Parse a CSV format levy export file.
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            LevyExportData object containing the parsed data
        """
        logger.info(f"Parsing CSV levy export file: {file_path}")
        
        records = []
        year = datetime.now().year
        
        try:
            # First, read a sample to analyze the file
            with open(file_path, 'r', encoding='utf-8-sig') as sample_file:
                sample = sample_file.read(1024)
                logger.info(f"CSV Sample: {sample[:100]}")
            
            # Map common header variations to standardized field names
            field_mappings = {
                'LEVY CODE': 'levy_cd',
                'LEVY_CODE': 'levy_cd',
                'CODE': 'levy_cd',
                'LEVY_CD': 'levy_cd',
                'LEVYCD': 'levy_cd',
                'LEVY': 'levy_cd',  # Sometimes just "LEVY" is used for the code
                'TAX_DISTRICT_ID': 'tax_district_id',
                'TAX DISTRICT': 'tax_district_id',
                'DISTRICT': 'tax_district_id',
                'DIST': 'tax_district_id',
                'RATE': 'levy_rate',
                'TAX RATE': 'levy_rate',
                'LEVY RATE': 'levy_rate',
                'LEVY_RATE': 'levy_rate',
                'AMOUNT': 'levy_amount',
                'LEVY AMOUNT': 'levy_amount',
                'LEVY_AMOUNT': 'levy_amount',
                'VALUE': 'assessed_value',
                'ASSESSED VALUE': 'assessed_value',
                'ASSESSED_VALUE': 'assessed_value',
                'LINKED': 'levy_cd_linked',
                'LINKED CODE': 'levy_cd_linked',
                'LINKED_CODE': 'levy_cd_linked',
                'LEVY_CD_LINKED': 'levy_cd_linked',
                'YEAR': 'year'
            }
            
            # Try to extract year from filename
            filename = file_path.name
            year_match = re.search(r'(\d{4})', filename)
            if year_match:
                try:
                    year_val = int(year_match.group(1))
                    if 2000 <= year_val <= 2100:  # Sanity check for year
                        year = year_val
                        logger.info(f"Found year in filename: {year}")
                except ValueError:
                    pass
            
            # Try CSV sniffing approach
            try:
                dialect = csv.Sniffer().sniff(sample)
                has_header = csv.Sniffer().has_header(sample)
                logger.info(f"CSV dialect detected: delimiter='{dialect.delimiter}', has_header={has_header}")
            except csv.Error:
                logger.warning("Could not determine CSV dialect, using default comma delimiter")
                dialect = csv.excel
                has_header = True if ',' in sample and len(sample.splitlines()) > 1 else False
            
            # Read the file with DictReader
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f, dialect=dialect)
                
                # Log field names if available
                if reader.fieldnames:
                    logger.info(f"CSV field names: {reader.fieldnames}")
                else:
                    logger.warning("No field names detected in CSV")
                
                # Create a mapping for this specific CSV's headers
                header_map = {}
                if reader.fieldnames:
                    for header in reader.fieldnames:
                        if header is None:
                            continue
                        upper_header = header.upper()
                        for key, value in field_mappings.items():
                            if key == upper_header or key in upper_header:
                                header_map[header] = value
                                break
                
                logger.info(f"CSV header mapping: {header_map}")
                
                row_count = 0
                # Process each row
                for row in reader:
                    row_count += 1
                    logger.debug(f"Processing CSV row {row_count}: {row}")
                    
                    # Skip empty rows
                    if not row or all(not value for value in row.values()):
                        continue
                    
                    record = {
                        'tax_district_id': None,
                        'levy_cd': None,
                        'levy_cd_linked': None,
                        'levy_rate': None,
                        'levy_amount': None,
                        'assessed_value': None,
                        'year': year,
                        'source': 'csv'
                    }
                    
                    # Map values using the header map
                    for original_header, value in row.items():
                        if original_header is None or not value:
                            continue
                            
                        mapped_field = header_map.get(original_header)
                        if mapped_field:
                            # Special handling for levy code which might include a linked code
                            if mapped_field == 'levy_cd' and value and '/' in value:
                                levy_cd, levy_cd_linked = value.split('/', 1)
                                record['levy_cd'] = levy_cd.strip()
                                record['levy_cd_linked'] = levy_cd_linked.strip()
                                if not record['tax_district_id']:
                                    record['tax_district_id'] = levy_cd.strip()
                            elif mapped_field == 'levy_cd' and value:
                                record['levy_cd'] = value.strip()
                                if not record['tax_district_id']:
                                    record['tax_district_id'] = value.strip()
                            # Convert numeric fields
                            elif mapped_field in ('levy_rate', 'levy_amount', 'assessed_value'):
                                try:
                                    # Remove any commas or currency symbols
                                    clean_value = re.sub(r'[^\d.-]', '', str(value))
                                    record[mapped_field] = float(clean_value) if clean_value else None
                                except (ValueError, TypeError):
                                    logger.warning(f"Could not convert {value} to float for {mapped_field}")
                            elif mapped_field == 'year' and value:
                                try:
                                    year_val = int(value)
                                    if 1900 <= year_val <= 2100:  # Sanity check for year
                                        record['year'] = year_val
                                except (ValueError, TypeError):
                                    pass
                            else:
                                record[mapped_field] = value
                    
                    # If we have a tax_district_id but no levy_cd, use tax_district_id as levy_cd
                    if record['tax_district_id'] and not record['levy_cd']:
                        record['levy_cd'] = record['tax_district_id']
                    
                    # If we have a levy_cd but no tax_district_id, use levy_cd as tax_district_id
                    if record['levy_cd'] and not record['tax_district_id']:
                        record['tax_district_id'] = record['levy_cd']
                    
                    # Only add record if we have a levy code or tax_district_id
                    if record['levy_cd'] or record['tax_district_id']:
                        records.append(record)
                        logger.debug(f"Added record: {record}")
            
            # If no records were found, try parsing without headers as a fallback
            if not records:
                logger.info("Attempting to parse CSV without headers")
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f, dialect=dialect)
                    first_row = True
                    row_count = 0
                    
                    for row in reader:
                        row_count += 1
                        # Skip header row
                        if first_row and has_header:
                            first_row = False
                            continue
                        
                        # Skip empty rows
                        if not row or all(not cell for cell in row):
                            continue
                        
                        # Try to extract data based on position
                        if len(row) >= 3:  # At minimum need district, levy code, and rate
                            tax_district_id = row[0].strip() if len(row) > 0 and row[0] else None
                            levy_cd = row[1].strip() if len(row) > 1 and row[1] else tax_district_id
                            levy_cd_linked = row[2].strip() if len(row) > 2 and row[2] else None
                            
                            try:
                                levy_rate = float(re.sub(r'[^\d.-]', '', row[3])) if len(row) > 3 and row[3] else None
                            except (ValueError, TypeError):
                                levy_rate = None
                                
                            try:
                                levy_amount = float(re.sub(r'[^\d.-]', '', row[4])) if len(row) > 4 and row[4] else None
                            except (ValueError, TypeError):
                                levy_amount = None
                            
                            try:
                                year_val = int(row[5]) if len(row) > 5 and row[5] and row[5].isdigit() else year
                                if not (1900 <= year_val <= 2100):  # Sanity check
                                    year_val = year
                            except (ValueError, TypeError):
                                year_val = year
                            
                            record = {
                                'tax_district_id': tax_district_id,
                                'levy_cd': levy_cd,
                                'levy_cd_linked': levy_cd_linked,
                                'levy_rate': levy_rate,
                                'levy_amount': levy_amount,
                                'year': year_val,
                                'source': 'csv'
                            }
                            
                            # Ensure we have either tax_district_id or levy_cd
                            if record['tax_district_id'] or record['levy_cd']:
                                records.append(record)
                                logger.debug(f"Added record from no-header parsing: {record}")
        
        except Exception as e:
            logger.error(f"Error parsing CSV file {file_path}: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return LevyExportData([], {'format': 'csv', 'year': year, 'error': str(e)})
        
        logger.info(f"Parsed {len(records)} records from CSV file")
        return LevyExportData([r for r in records if isinstance(r, dict)], {'format': 'csv', 'year': year})
        
    @classmethod
    def _parse_json(cls, file_path: Path) -> LevyExportData:
        """
        Parse a JSON format levy export file.
        
        Args:
            file_path: Path to the JSON file
            
        Returns:
            LevyExportData object containing the parsed data
        """
        logger.info(f"Parsing JSON levy export file: {file_path}")
        
        import json
        records = []
        year = datetime.now().year
        
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                data = json.load(f)
            
            # Try to extract year from metadata if available
            if isinstance(data, dict):
                metadata = data.get('metadata', {})
                if metadata and 'year' in metadata:
                    try:
                        year = int(metadata['year'])
                        logger.info(f"Found year in JSON metadata: {year}")
                    except (ValueError, TypeError):
                        pass
                
                # Look for records array
                levy_records = data.get('records', [])
                if not levy_records and 'levies' in data:
                    levy_records = data.get('levies', [])
                if not levy_records and 'districts' in data:
                    levy_records = data.get('districts', [])
            elif isinstance(data, list):
                # Assume the array is the records
                levy_records = data
            else:
                levy_records = []
                logger.warning(f"Unexpected JSON structure in {file_path}")
            
            # Process records
            for item in levy_records:
                if not isinstance(item, dict):
                    continue
                    
                # Extract levy code
                levy_cd = None
                for key in ('levy_cd', 'levy_code', 'code', 'levy', 'districtId', 'district_id'):
                    if key in item and item[key]:
                        levy_cd = str(item[key]).strip()
                        break
                
                if not levy_cd:
                    continue
                
                # Extract other fields
                levy_cd_linked = None
                if '/' in levy_cd:
                    levy_cd, levy_cd_linked = levy_cd.split('/', 1)
                
                # Try to extract rate, amount and value
                rate = None
                for key in ('rate', 'levy_rate', 'tax_rate'):
                    if key in item and item[key] is not None:
                        try:
                            rate = float(item[key])
                            break
                        except (ValueError, TypeError):
                            pass
                
                amount = None
                for key in ('amount', 'levy_amount', 'levy'):
                    if key in item and item[key] is not None:
                        try:
                            amount = float(item[key])
                            break
                        except (ValueError, TypeError):
                            pass
                
                value = None
                for key in ('value', 'assessed_value', 'assessed'):
                    if key in item and item[key] is not None:
                        try:
                            value = float(item[key])
                            break
                        except (ValueError, TypeError):
                            pass
                
                # Extract year from record if available
                record_year = year
                for key in ('year', 'tax_year', 'levy_year'):
                    if key in item and item[key] is not None:
                        try:
                            year_val = int(item[key])
                            if 1900 <= year_val <= 2100:  # Sanity check
                                record_year = year_val
                                break
                        except (ValueError, TypeError):
                            pass
                
                record = {
                    'tax_district_id': levy_cd,
                    'levy_cd': levy_cd,
                    'levy_cd_linked': levy_cd_linked,
                    'levy_rate': rate,
                    'levy_amount': amount,
                    'assessed_value': value,
                    'year': record_year,
                    'source': 'json'
                }
                records.append(record)
        
        except Exception as e:
            logger.error(f"Error parsing JSON file {file_path}: {str(e)}")
            raise
        
        logger.info(f"Parsed {len(records)} records from JSON file")
        return LevyExportData(records, {'format': 'json', 'year': year})
    
    @classmethod
    def _parse_xml(cls, file_path: Path) -> LevyExportData:
        """
        Parse an XML format levy export file.
        
        Args:
            file_path: Path to the XML file
            
        Returns:
            LevyExportData object containing the parsed data
        """
        logger.info(f"Parsing XML levy export file: {file_path}")
        
        try:
            import xml.etree.ElementTree as ET
            
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Extract year from metadata if available
            year = datetime.now().year
            year_elements = root.findall('.//year') or root.findall('.//Year')
            if year_elements:
                try:
                    year = int(year_elements[0].text)
                    logger.info(f"Found year in XML: {year}")
                except (ValueError, TypeError):
                    pass
            
            # Parse levy records
            records = []
            
            # Try different possible structures
            levy_elements = (
                root.findall('.//levy') or 
                root.findall('.//Levy') or 
                root.findall('.//levy_record') or
                root.findall('.//LevyRecord')
            )
            
            if not levy_elements:
                logger.warning("No levy records found in XML structure")
            
            for levy_elem in levy_elements:
                # Extract code
                code_elem = (
                    levy_elem.find('./code') or
                    levy_elem.find('./Code') or
                    levy_elem.find('./levy_code') or
                    levy_elem.find('./LevyCode')
                )
                
                if code_elem is None or not code_elem.text:
                    continue
                
                levy_cd = code_elem.text.strip()
                
                # Extract rate
                rate_elem = (
                    levy_elem.find('./rate') or
                    levy_elem.find('./Rate')
                )
                rate = None
                if rate_elem is not None and rate_elem.text:
                    try:
                        rate = float(rate_elem.text)
                    except (ValueError, TypeError):
                        pass
                
                # Extract levy
                levy_amount_elem = (
                    levy_elem.find('./amount') or
                    levy_elem.find('./Amount') or
                    levy_elem.find('./levy_amount') or
                    levy_elem.find('./LevyAmount')
                )
                levy_amount = None
                if levy_amount_elem is not None and levy_amount_elem.text:
                    try:
                        levy_amount = float(levy_amount_elem.text)
                    except (ValueError, TypeError):
                        pass
                
                # Extract value
                value_elem = (
                    levy_elem.find('./value') or
                    levy_elem.find('./Value') or
                    levy_elem.find('./assessed_value') or
                    levy_elem.find('./AssessedValue')
                )
                value = None
                if value_elem is not None and value_elem.text:
                    try:
                        value = float(value_elem.text)
                    except (ValueError, TypeError):
                        pass
                
                # Check for linked levy code
                levy_cd_linked = None
                if '/' in levy_cd:
                    levy_cd, levy_cd_linked = levy_cd.split('/', 1)
                
                record = {
                    'tax_district_id': levy_cd,
                    'levy_cd': levy_cd,
                    'levy_cd_linked': levy_cd_linked,
                    'levy_rate': rate,
                    'levy_amount': levy_amount,
                    'assessed_value': value,
                    'year': year,
                    'source': 'xml'
                }
                records.append(record)
        
        except Exception as e:
            logger.error(f"Error parsing XML file {file_path}: {str(e)}")
            raise
        
        logger.info(f"Parsed {len(records)} records from XML file")
        return LevyExportData(records, {'format': 'xml', 'year': year})
        
    @classmethod
    def create_template(cls, format_type: LevyExportFormat, include_sample_data: bool = True) -> Union[str, bytes]:
        """
        Create a template file for levy exports.
        
        Args:
            format_type: Format type to create template for
            include_sample_data: Whether to include sample data in the template
            
        Returns:
            File content as string or bytes depending on format
        """
        year = datetime.now().year
        sample_records = []
        
        if include_sample_data:
            # Create sample data
            sample_records = [
                {
                    'tax_district_id': '001',
                    'levy_cd': '001-001',
                    'levy_cd_linked': '001-002',
                    'levy_rate': 5.5000,
                    'levy_amount': 550000.00,
                    'assessed_value': 10000000.00,
                    'year': year
                },
                {
                    'tax_district_id': '002',
                    'levy_cd': '002-001',
                    'levy_cd_linked': '002-002',
                    'levy_rate': 6.2500,
                    'levy_amount': 312500.00,
                    'assessed_value': 5000000.00,
                    'year': year
                },
                {
                    'tax_district_id': '003',
                    'levy_cd': '003-001',
                    'levy_cd_linked': '003-002',
                    'levy_rate': 4.8750,
                    'levy_amount': 243750.00,
                    'assessed_value': 5000000.00,
                    'year': year
                }
            ]
        else:
            # Create empty template with just column headers
            sample_records = [
                {
                    'tax_district_id': '',
                    'levy_cd': '',
                    'levy_cd_linked': '',
                    'levy_rate': None,
                    'levy_amount': None,
                    'assessed_value': None,
                    'year': year
                }
            ]
        
        # Create template based on format type
        if format_type == LevyExportFormat.CSV:
            # Create CSV template
            output = io.StringIO()
            fieldnames = ['tax_district_id', 'levy_cd', 'levy_cd_linked', 'levy_rate', 'levy_amount', 'assessed_value', 'year']
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            for record in sample_records:
                writer.writerow(record)
            return output.getvalue()
            
        elif format_type == LevyExportFormat.JSON:
            # Create JSON template
            data = {
                'metadata': {
                    'year': year,
                    'created_at': datetime.now().isoformat(),
                    'format_version': '1.0'
                },
                'records': sample_records
            }
            return json.dumps(data, indent=2)
            
        elif format_type == LevyExportFormat.XLSX:
            # Create Excel template
            output = io.BytesIO()
            df = pd.DataFrame(sample_records)
            writer = pd.ExcelWriter(output, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Levy Data')
            
            # Add a documentation sheet
            doc_data = {
                'Field': ['tax_district_id', 'levy_cd', 'levy_cd_linked', 'levy_rate', 'levy_amount', 'assessed_value', 'year'],
                'Description': [
                    'Tax district identifier',
                    'Primary levy code',
                    'Linked levy code (optional)',
                    'Levy rate per $1000 of assessed value',
                    'Total levy amount in dollars',
                    'Total assessed value in dollars',
                    'Tax year'
                ],
                'Example': [
                    '001',
                    '001-001',
                    '001-002',
                    '5.5000',
                    '550000.00',
                    '10000000.00',
                    str(year)
                ]
            }
            doc_df = pd.DataFrame(doc_data)
            doc_df.to_excel(writer, index=False, sheet_name='Documentation')
            
            writer.close()
            return output.getvalue()
            
        elif format_type == LevyExportFormat.XML:
            # Create XML template
            root = f"""<?xml version="1.0" encoding="UTF-8"?>
<levy_export>
    <metadata>
        <year>{year}</year>
        <created_at>{datetime.now().isoformat()}</created_at>
        <format_version>1.0</format_version>
    </metadata>
    <records>
"""
            
            for record in sample_records:
                root += f"""        <record>
            <tax_district_id>{record['tax_district_id']}</tax_district_id>
            <levy_cd>{record['levy_cd']}</levy_cd>
            <levy_cd_linked>{record['levy_cd_linked']}</levy_cd_linked>
            <levy_rate>{record['levy_rate']}</levy_rate>
            <levy_amount>{record['levy_amount']}</levy_amount>
            <assessed_value>{record['assessed_value']}</assessed_value>
            <year>{record['year']}</year>
        </record>
"""
            
            root += """    </records>
</levy_export>
"""
            return root
            
        elif format_type == LevyExportFormat.TXT:
            # Create TXT template (fixed width format)
            output = io.StringIO()
            output.write(f"LEVY EXPORT DATA - YEAR {year}\n")
            output.write("-" * 80 + "\n")
            output.write("LEVY CODE      RATE        LEVY        VALUE\n")
            output.write("-" * 80 + "\n")
            
            for record in sample_records:
                levy_cd = f"{record['levy_cd']}"
                if record['levy_cd_linked']:
                    levy_cd += f"/{record['levy_cd_linked']}"
                    
                rate = record['levy_rate'] if record['levy_rate'] is not None else 0
                levy = record['levy_amount'] if record['levy_amount'] is not None else 0
                value = record['assessed_value'] if record['assessed_value'] is not None else 0
                
                output.write(f"{levy_cd:<15} {rate:<11.4f} {levy:<11.2f} {value:<15.2f}\n")
                
            return output.getvalue()
        
        else:
            raise ValueError(f"Unsupported template format: {format_type}")